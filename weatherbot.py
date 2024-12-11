import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tabulate import tabulate
import pandas as pd
import datetime
import re
import random
import numpy as np

from selenium.common.exceptions import TimeoutException

ufo_airport = pd.read_excel('ufos_airports.xlsx')
ufo_airport['date'] = pd.to_datetime(ufo_airport['date'], format='%d/%m/%Y')
# airport_codes = np.unique(ufo_airport['icao'].values)

# years_all = list(range(2003, 2014)) # originally 1965
# months = range(1, 13)

# # Extract weather data from the website
# def extract_weather_data(driver, airport_code, year, month):
#     url = f"https://www.wunderground.com/history/monthly/{airport_code}/date/{year}-{month}"
#     driver.get(url)

#     try:
#         WebDriverWait(driver, 6).until(
#             EC.presence_of_element_located((By.XPATH, '//table[@class="days ng-star-inserted"][1]'))
#         )
#     except TimeoutException:
#         print(f"Timed out waiting for the table to load for {year}-{month}. Skipping this month.")
#         return None, None
    
#     city_name = driver.find_element(By.XPATH, '//h1[@_ngcontent-app-root-c176]/span[@_ngcontent-app-root-c176]').text
#     print(city_name)

#     headers = []
#     data = []

#     # Scrape the table headers
#     for th in driver.find_elements(By.XPATH, '//table[@class="days ng-star-inserted"][1]/tbody/tr[1]/td'):
#         headers.append(th.text)

#     # Scrape the table rows
#     for row in driver.find_elements(By.XPATH, '//table[@class="days ng-star-inserted"][1]/tbody/tr[position() > 1]'):
#         rowData = [airport_code, city_name]

#         firstCell = row.find_element(By.XPATH, './td[1]')
#         date_str = f"{firstCell.text}-{month}-{year}"
#         rowData.append(date_str)
#         for td in row.find_elements(By.XPATH, './td[position() > 1]'):
#             rowData.append(td.text)
#         data.append(rowData)

#     return headers, data

# # Write the scraped data to Excel
# def write_printed_output_to_excel(sheet, printed_output, airport_code, year, month, row_offset, first_month):
#     for row_num, row in enumerate(printed_output.split('\n'), 1):
#         if row.startswith("Date") or row.startswith("-----"):
#             continue

#         airport_cell = sheet.cell(row=row_num + row_offset, column=1)
#         airport_cell.value = airport_code
        
#         for col_num, cell_value in enumerate(row.split(), 2):
#             cell = sheet.cell(row=row_num + row_offset, column=col_num)
#             if col_num == 2 and row_num > 1:
#                 day_values = re.findall(r'\d+', cell_value)
#                 if day_values:
#                     day = day_values[0]
#                     date_str = f"{year}-{month}-{day}"
#                     cell.value = datetime.datetime.strptime(date_str, "%Y-%m-%d")
#                     cell.number_format = 'dd-mmm-yyyy'
#                 else:
#                     continue
#             else:
#                 cell.value = cell_value
#             if row_num == 1 and first_month:
#                 cell.font = openpyxl.styles.Font(bold=True)

#     for col_num in range(1, len(headers) + 2):
#         column_letter = get_column_letter(col_num)
#         max_length = max(len(str(cell.value)) for cell in sheet[column_letter])
#         sheet.column_dimensions[column_letter].width = max_length + 2

#     return row_offset + row_num - 1

# # Main script
# options = webdriver.ChromeOptions()
# options.add_argument('--headless=new')
# driver = webdriver.Chrome(options=options)
# wb = Workbook()
# ws = wb.active
# row_offset = 0
# first_month = True

# for airport_code in airport_codes:
#     years = random.sample(years_all, k=10) # originally k=10
#     print(airport_codes)
#     for year in years:
#         random_months = random.sample(list(months), k=3) # originally k=3

#         for month in random_months:
#             print(f"Scraping data for {airport_code} - {year}-{month}")
#             headers, data = extract_weather_data(driver, airport_code, year, month)

#             if headers is None and data is None:
#                 print(f"No data found for {airport_code} - {year}-{month}. Skipping this month.")
#                 continue

#             printed_output = tabulate(data, headers=headers)
#             print(printed_output)
#             print("\n")

#             if headers is not None and data is not None:
#                 row_offset = write_printed_output_to_excel(ws, printed_output, airport_code, year, month, row_offset, first_month)

#             first_month = False

# for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
#     date_cell = row[0]
#     date_cell.number_format = 'dd-mmm-yyyy'

# wb.save("weather_data_combined2.xlsx")
# print("All data saved to weather_data_combined.xlsx\n")

# driver.quit()
# left_on=['icao', 'datetime'], right_on=ufo_airport.iloc[:, [0, 1]]
weather = pd.read_excel('weather_data_combined2.xlsx')
weather['date'] = pd.to_datetime(weather['date'], format='%d/%m/%Y')
table_final = pd.merge(ufo_airport, weather, 'outer', on=['icao', 'date'])
table_final.to_excel('merged_weather_ufo.xlsx', index=False)